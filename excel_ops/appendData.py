import pandas as pd
from openpyxl import load_workbook

def append(file_path, name, ph_num, email, requirement, communication):
    try:
        if not any([name, ph_num, email, requirement, communication]):
            print("Skipping empty data")
            return

        df = pd.DataFrame.from_records([{
            "name": name or "",
            "phone number": ph_num or "",
            "email": email or "",
            "requirement": requirement or "",
            "communication": communication or ""
        }])

        print("DF:\n", df)

        from openpyxl import load_workbook
        book = load_workbook(file_path)
        sheet = book["Sheet1"]
        start_row = sheet.max_row
        book.close()

        with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            df.to_excel(
                writer,
                sheet_name="Sheet1",
                startrow=start_row,
                index=False,
                header=False
            )

        print("Append successful")

        

    except Exception as e:
        print("Error in appending data to excel:", e)